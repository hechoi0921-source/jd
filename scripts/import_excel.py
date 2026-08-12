"""Convert the approved workbook to browser-ready JSON while preserving source rows."""
import json,sys
from pathlib import Path
from openpyxl import load_workbook
src=Path(sys.argv[1]) if len(sys.argv)>1 else Path.home()/"OneDrive"/"문서"/"seoyeon_stdjob_full_260708.xlsx"
out=Path(__file__).parents[1]/"public"/"data"/"jobs.json";out.parent.mkdir(parents=True,exist_ok=True)
wb=load_workbook(src,data_only=True,read_only=True)
def rows(name):
 ws=wb[name];it=ws.iter_rows(values_only=True);head=[str(x).strip() for x in next(it)]
 return [(i,dict(zip(head,r))) for i,r in enumerate(it,2) if any(v is not None for v in r)]
jd=rows('job_description');ta=rows('task_activity');sk=rows('skill')
jobs=[]
for source_row,r in jd:
 jid=int(r['No.']); job={'id':jid,'source_row':source_row,'job_group':r['직군'],'job_series':r['직렬'],'name':r['직무'],'description':r['정의(Description)'] or '', 'mission':r['목적(Mission)'] or '', 'companies':[c for c in ('서연','서연이화','서연탑메탈') if r.get(c)=='O'],'tasks':[],'skills':[]}
 taskmap={}
 for rr,x in ta:
  if int(x['No.'])!=jid: continue
  name=x.get('Task')
  if not name: continue
  if name not in taskmap:
   t={'id':f'j{jid}-t{len(taskmap)+1}','order':len(taskmap)+1,'name':name,'source_row':rr,'activities':[]};taskmap[name]=t;job['tasks'].append(t)
  if x.get('Activity'): taskmap[name]['activities'].append({'id':f'j{jid}-a{rr}','text':x['Activity'],'source_row':rr})
 for rr,x in sk:
  if int(x['No.'])==jid: job['skills'].append({'id':f'j{jid}-s{rr}','name':x['skill_name'],'hard_soft':x['hard_soft'],'ksao':x['KSAO'],'description':x['skill_description'] or '', 'related_task':x['related_task'] or '', 'source_row':rr})
 jobs.append(job)
assert (len(jobs),sum(len(j['tasks']) for j in jobs),sum(len(t['activities']) for j in jobs for t in j['tasks']),sum(len(j['skills']) for j in jobs))==(90,365,1100,988)
out.write_text(json.dumps(jobs,ensure_ascii=False,separators=(',',':')),encoding='utf-8');print(out)
